#%% zwraca daną komórkę 
#modyf jp: dwa sposoby określania zwrotu komórki
import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
cell_0 = sheet['a1']  #col,row
cell = sheet.cell(1, 2)  #row,col
print(cell_0.value)
print(cell.value)

#%%  wb to workbook
#sprawdza ilość wierszy w arkuszu w którym są dane

import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1'] 
#cell = sheet.cell(1, 1)
print(sheet.max_row)

#%%  wynik powinien byc w pionie liczby 1,2,3,4
#badamy range od wiersza 1 do 4, a ponieważ range 
#nie obejmuje tej ostatniej wartości, to trzeba 
#dodać 1: sheet.max_row + 1
import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (1, sheet.max_row + 1):
   print(row)

# %% przeszukujemy od wiersza nr 2 (bo w 1 jest string:nazwa price)
# i do max. ilości wierszy, czyli 4 (w których są dane)
#i w kolumnie nr 3 (tu są wartości liczbowe cen)
import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   print(cell.value)

# %% korygujemy cenę, tj. w nowej kolumnie nr 4 
# przemnażamy wartości przez 0.9 z kolumny nr 3
# tworzymy nowy plik transactions2.xlsx 
#  ze skorygowaną ceną
import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price
   
wb.save('transactions2.xlsx')

# %% modyfikacja jp- zapisuje skorygowaną cenę 
# w kolumnie 4 tego samego arkusza 
# (nie trzeba dodawać save, tj. zapis tego pliku, bo też działa)
import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price

#wb.save('transactions.xlsx')

# %% modyfikacja jp: dodałam print skorygowanych cen
# i zapisałam w pliku transactions3.xlsx

import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
    print(corrected_price_cell.value)

wb.save('transactions3.xlsx')

# %% dodatkowo tworzenie wykresu - wymaga dodania klas 
# BartChart i Reference

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell = sheet['a1']
#cell = sheet.cell(1, 1)

for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price

values = Reference(sheet, 
         min_row=2, 
         max_row=sheet.max_row,
         min_col=4,
         max_col=4)    

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')

#%% CAŁOSC  z 1 plikiem
#wb to workbook
#przeszukujemy od wiersza nr 2 (bo w 1 jest string:nazwa price)
#  i do max. ilości wierszy, czyli 4 (w których są dane)
#  i w kolumnie nr 3 (tu są wartości liczbowe cen)

#range od wiersza 1 do 4, a ponieważ range 
#  nie obejmuje tej ostatniej wartości, to trzeba 
#  dodać 1: sheet.max_row + 1

#przeszukujemy od wiersza nr 2 (bo w 1 jest string:nazwa price)
#  i do max. ilości wierszy, czyli 4 (w których są dane)
#  i w kolumnie nr 3 (tu są wartości liczbowe cen)

#korygujemy cenę, tj. w nowej kolumnie nr 4 
#  przemnażamy wartości przez 0.9 z kolumny nr 3
#  tworzymy nowy plik transactions2.xlsx 
#  ze skorygowaną ceną
#dodałam print skorygowanych cen
#dodatkowo tworzenie wykresu - wymaga dodania klas 
#  BartChart i Reference

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']
#cell_0 = sheet['a1']         #col,row
#cell_1 = sheet.cell(1, 1)    #row,col
#print(cell_0.value)
#print(cell_1.value)
#print(sheet.max_row) #zwraca ilość wierszy

#for row in range (1, sheet.max_row + 1):
   #print(row)


for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price
   print(cell.value)
   print(corrected_price_cell.value)

values = Reference(sheet, 
         min_row=2, 
         max_row=sheet.max_row,
         min_col=4,
         max_col=4)    

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')


# %%  kod -dodana funckcja def process_worbook
# zmiana na filename (dot.rożnych exceli)
#zapis w tym samym pliku
#nie działa-może trzeba dodac ścieżkę do folderu z excelami

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):


   wb = xl.load_workbook(filename)
   sheet = wb['Arkusz1']
   
   for row in range (2, sheet.max_row + 1):
      cell = sheet.cell(row, 3)
      corrected_price = cell.value * 0.9
      corrected_price_cell = sheet.cell(row, 4)
      corrected_price_cell.value = corrected_price

   values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)    

   chart = BarChart()
   chart.add_data(values)
   sheet.add_chart(chart, 'e2')

   wb.save(filename)


# %% usunac
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Arkusz1']


for row in range (2, sheet.max_row + 1):
   cell = sheet.cell(row, 3)
   corrected_price = cell.value * 0.9
   corrected_price_cell = sheet.cell(row, 4)
   corrected_price_cell.value = corrected_price
   print(cell.value)
   print(corrected_price_cell.value)

for row in range (1, sheet.min_row + 1):
   cell_1 = sheet.cell(row, 4)
   new_price = cell.value()
   new_price_cell = wb.cell(row=1, column=4, value=2)
   new_price_cell.value = new_price

values = Reference(sheet, 
         min_row=2, 
         max_row=sheet.max_row,
         min_col=4,
         max_col=4)    

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')

# %%
